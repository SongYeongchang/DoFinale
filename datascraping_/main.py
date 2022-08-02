# -*- coding: utf-8 -*-

from selenium import webdriver
import chromedriver_autoinstaller
from selenium.webdriver.common.by import By
import os, time

import pandas as pd
from openpyxl import load_workbook, Workbook

import requests, json, random
from bs4 import BeautifulSoup
from urllib import request

from geopy.geocoders import Nominatim



def location_keywords(excel_file_dir):
    '''
    전국행정동 리스트 엑셀 파일을 읽고
    "대분류 + 구" 추출하기
    ec) 서울 강남구

    :param excel_file_dir: 엑셀 파일
    :return: 지역명
    '''
    read_xlsx = load_workbook(excel_file_dir)
    read_sheet = read_xlsx.active

    Main_Category = [] # 대분류(서울특별시, 광주광역시, 경기도 등)
    Sub_Category = [] # 구
    Locations = [] # 검색 지역명

    name_col_A = read_sheet['A'] # 대분류
    name_col_C = read_sheet['C'] # 구
    name_col_D = read_sheet['D'] # 동/면/리


    # Main_Category 정리
    for i, cell in enumerate(name_col_A):
        if cell.value: # 대분류 값이 존재할 경우 추가
            Main_Category.append(cell.value)

    Main_Category.remove('대분류')
    Main_Category.remove('전국행정동리스트')

    # 특별시, 광역시 제거
    for i, itm in enumerate(Main_Category):
        if '특별시' in itm:
            Main_Category[i] = Main_Category[i].replace('특별시','')
        elif '광역시' in itm:
            Main_Category[i] = Main_Category[i].replace('광역시','')

    # Sub_Category 정리
    for i, cell in enumerate(name_col_C):
        if cell.value: # 구 값이 존재할 경우 추가
            Sub_Category.append(cell.value)
        elif (not cell.value and name_col_D[i].value): # 구 값이 없고 동시에 동/면/리 값 있는 경우, 동/면/리 값 추가
            Sub_Category.append(name_col_D[i].value)
        # else: # 구, 동/면/리 없는 경우 종료
        #     break
    Sub_Category.remove('구')

    # 지역명
    for i in range(len(Main_Category)):
        Locations.append(Main_Category[i] + ' ' + Sub_Category[i])

    # # 최종 결과 확인
    # print(Main_Category)
    # print('길이=',len(Main_Category))
    # set_Main_Category=set(Main_Category)
    # print(set_Main_Category)
    # print(len(set_Main_Category))
    # print('---------------------------------')
    # print(Sub_Category)
    # print('길이=',len(Sub_Category))
    # print('---------------------------------')
    # print(Locations)

    return Locations

def search_result_2_json(search_word):
    '''
    네이버 지도에서 각 지역 (대분류 + 구 + search_word + 병원) 검색 후
    결과 엑셀로 저장

    네이버 지도 검색 후 requests.get(url)로 데이터 읽을 때
    필요한 정보가 json 형태로 있음.

    키워드 검색 후
    필요한 정보가 담긴 json 데이터를
    ./data 디렉토리 안에 저장

    json 형식의 데이터 최대 15개 까지 검색 결과가 들어감. 해당 부분 알 수 있게끔 데이터 저장

    :param search_word:
    :return:
    '''
    # 검색어 만들기
    excel_file_dir = '전국행정동리스트_.xlsx'
    Locations = location_keywords(excel_file_dir)
    keyword = search_word
    keywords = [] # 검색어들을 저장할 리스트
    for i, location in enumerate(Locations):
        keywords.append(location + ' ' + keyword + '병원') # 지역(대분류 + 구) + 키워드(탈모,모발,두피) + 병원

    alpha = 0 # 추가 대기 시간, 404에러 후 다시 접속할 때까지 기다리는 시간
    # http://json.parser.online.fr/
    # https://jsonformatter.org/json-parser
    # https://jsonparser.org/
    for i, itm in enumerate(keywords):
        url = "https://m.map.naver.com/search2/search.naver?query=" + itm
        html = requests.get(url)
        # requests.get(url) 결과가 404일 경우
        while html.status_code == 404:
            html = requests.get(url)
            print(i, itm, '진행. 404에러, 재시도중..')
            time.sleep(random.randrange(1, 60) + alpha)
            alpha = alpha + 30 # 대기 시간 추가
        try:
            # 필요한 json 데이터 추출
            html = html.text
            html = html.split('var searchResult = ')[1]
            html = html.split('var searchService =')[0]
            html = html.replace(';', '').strip()

            print(itm, i, '번째 병원 개수 len>>>>>', len(json.loads(html)['site']['list']))
            if len(json.loads(html)['site']['list']) > 14:
                print(i, itm, '최대치 넘어감')
                # json 저장
                file_path = './data' + str(i) + itm + '최대치(15) 넘어감.json'  # https://jsikim1.tistory.com/221
                with open(file_path, 'w', encoding='utf-8') as file:
                    json.dump(html, file, indent="\t")
            else:
                # json 저장
                file_path = './data' + str(i) + itm + '.json'  # https://jsikim1.tistory.com/221
                with open(file_path, 'w', encoding='utf-8') as file:
                    json.dump(html, file, indent="\t")


        except: # 검색결과 없을 경우
            print(i, itm, '오류 발생. 검색결과 없음 예상')
            html = '검색결과 없음'

            # json 저장
            file_path = './data' + str(i) + itm + '_검색결과 없음.json'  # https://jsikim1.tistory.com/221
            with open(file_path, 'w', encoding='utf-8') as file:
                json.dump(html, file, indent="\t")

            # https://m.map.naver.com/search2/search.naver?query=서울 중랑구 탈모병원
            # 6번째 결과 없음
        alpha = 0  # 추가 대기 시간 초기화

def get_lat_and_log(address):
    '''
    주소 입력 시 위경도 값 출력하는 함수

    https://wonhwa.tistory.com/29
    :param address: 주소값
    :return: 위도값, 경도값
    '''
    try:
        geo_local = Nominatim(user_agent='South Korea')

        while(not geo_local.geocode(address)):
            # 마지막 어절 하나씩 제거
            address = address[:-len(address.split(' ')[-1])-1]
        geo=geo_local.geocode(address)

        return geo.latitude, geo.longitude # 위도, 경도 순으로 반환
    except:
        print(address,'에러!!!')
        return (0,0)

def dedamo_links():
    '''
    대다모 병원 링크 추출해서 엑셀 파일("hospitals_link_list.xlsx")로 저장
    '''
# ----- [ 대다모 병원 개별 링크 추출] -------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "병원정보"  # 엑셀 시트명 변경
    ws.append(["병원 링크 리스트"])

    for j in range(1,11): # 페이지 1 ~ 10 까지
        tmp = []
        url = "https://daedamo.com/new/bbs/board.php?bo_table=hlist?&page=" + str(j)
        html = requests.get(url).text
        #print('>>', html)
        bs = BeautifulSoup(html, "html5lib")
        for i in bs.select(".wr_id_first > a"): # 각 페이지에 모든 병원 링크
            print(i['href'])
            tmp.append(i['href']) # 리스트에 병원 링크 추가
            ws.append(tmp) # 엑셀에 병원 링크 추가된 리스트 추가
            tmp=[] # 리스트 초기화
        # print(len(bs.select(".wr_id_first > a")))
        print('-------------------')
        print(j)

    wb.save("hospitals_link_list.xlsx")

def dedamo_info(url):
    '''
    대다모 병원 정보
    :param url: 병원 각 개별 링크
    :return: 병원 링크, 병원명, 병원 주소, 병원 홈페이지, 병원 전화번호
    '''
    html = requests.get(url).text
    #print(html) # subject
    bs = BeautifulSoup(html, "html5lib")
    hos_name=bs.select_one(".subject").text
    hos_address=bs.find(class_="fa fa-map-marker").parent.text.strip()
    try:
        hos_homepage=bs.select(".hos_area > div.home > a")[0]['href']
    except:
        hos_homepage = '홈페이지 없음'
    hos_tel=bs.find(class_="fa fa-phone").parent.text.replace('전화상담','').replace('입니다','').replace('.','').strip()


    print("병원명:", hos_name)
    print("병원 주소:",  hos_address)
    print("병원 홈페이지:", hos_homepage)
    print("병원 전화번호:", hos_tel)

    res = []
    res.append(url)
    res.append(hos_name)
    res.append(hos_address)
    res.append(hos_homepage)
    res.append(hos_tel)

    return res

def excel_2_json_products(excel_file, result_path='./', products_json_form='./products_form.json', img_save_path='static/images/products/'):
    '''
    엑셀에 있는 데이터를 읽고
    json 형식 파일에 맞춰 값들 입력 후
    products.json 파일을 생성 해주는 함수

    excel_file: 엑셀 파일
    result_path: json 결과물 저장 경로
    products_json_form: json 형식 파일
    img_save_path: 제품 이미지 저장 경로

    :return: 성공 시 1, 실패 시 0
    '''
    print('제품 엑셀 파일 json으로 변환 시작')
    try:
        # 제품 엑셀 파일 df으로 읽기
        #df_products = pd.read_excel('추천_제품_리스트.xlsx', engine='openpyxl')
        df_products = pd.read_excel(excel_file, engine='openpyxl')
        # 데이터 프레임 테스트
        '''
        # print(df_products)
        # print('='*100)
        # print(len(df_products))
        # print(df_products.loc[0])
        # print('='*100)
        # print(df_products.loc[0]['scalp_type']=='양호' and df_products.loc[0]['category'] == '브랜드평판지수')
        # print(df_products.loc[0]['product_name'])
    
    
        # for i, itm in df_products.iterrows():
        #     print(itm)
        #     print(itm['scalp_type'])
        #     print(i%3,'*'*1000)
        '''

        scalp_types = ['양호', '건성', '지성', '민감성', '지루성', '염증성', '비듬성', '탈모성']
        categories = ['브랜드평판지수', '의사추천', '화해 랭킹순']

        # json(형식) 파일 열기
        #with open('./products_form.json', 'r', encoding="UTF-8") as file:
        with open(products_json_form, 'r', encoding="UTF-8") as file:
            data = json.load(file)
            # print(data) # json 형식 확인
            for scalp_type in scalp_types:  # ['양호','건성','지성','민감성','지루성','염증성','비듬성','탈모성']
                for category in categories:  # ['브랜드평판지수','의사추천','화해 랭킹순']
                    for j in range(0, 3):  # 제품 개수 3개 (0,1,2)
                        for i, itm in df_products.iterrows():  # 엑셀 데이터(df) 하나씩 비교
                            # scalp_type, category, 제품 개수 3개 (0,1,2) << 조건 3개 같은지 확인 후 데이터 입력
                            if (scalp_type == itm['scalp_type'] and category == itm['category'] and j == i % 3):
                                data['scalp_type'][scalp_type][category][j]['product_name'] = itm['product_name']
                                data['scalp_type'][scalp_type][category][j]['company'] = itm['company']
                                data['scalp_type'][scalp_type][category][j]['price'] = itm['price']
                                # data['scalp_type'][scalp_type][category][j]['image']=itm['image']
                                data['scalp_type'][scalp_type][category][j]['line'] = itm['line']
                                data['scalp_type'][scalp_type][category][j]['remarks'] = itm['remarks']

                                # # 이미지 저장 (img_save_path에 제품이름.jpg로 저장, json파일에 제품이름.jpg로 저장)
                                # print(itm['image'])
                                try:
                                    data['scalp_type'][scalp_type][category][j]['image'] = itm['product_name'] + '.jpg'
                                    #request.urlretrieve(itm['image'],'static/images/products/' + itm['product_name'] + '.jpg')
                                    request.urlretrieve(itm['image'],img_save_path + itm['product_name'] + '.jpg')
                                except Exception as e:
                                    # # request.urlretrieve 사용 시 403 에러 (아래 링크 참고)
                                    # # https://blog.naver.com/PostView.nhn?blogId=jinyuri303&logNo=222268855640&categoryNo=39&parentCategoryNo=0&viewDate=&currentPage=1&postListTopCurrentPage=1&from=postView
                                    # print(e)
                                    # print('에러 발생 이미지 url>', itm['image'])
                                    # print(itm['product_name'] + '.jpg')
                                    url_img = itm['image']
                                    with open(img_save_path + itm['product_name'] + '.jpg', "wb") as outfile:
                                        outfile.write(requests.get(url_img).content)
                                        #print('에러 처리 완료')

                                # 입력된 데이터값 확인
                                '''
                                # print(data['scalp_type'][scalp_type][category][j]['product_name'])
                                # print(data['scalp_type'][scalp_type][category][j]['company'])
                                # print(data['scalp_type'][scalp_type][category][j]['price'])
                                # print(data['scalp_type'][scalp_type][category][j]['image'])
                                # print(data['scalp_type'][scalp_type][category][j]['line'])
                                # print(data['scalp_type'][scalp_type][category][j]['remarks'])
                                '''

            # json 파일 저장
            with open(result_path+'products.json', 'w', encoding='utf-8') as file:
                json.dump(data, file)
                # json.dump(data, file, indent="\t")

        # 저장한 json 파일 읽고 출력
        with open(result_path+'products.json', 'r', encoding="UTF-8") as file:
            data = json.load(file)
            #print(data)  # 출력된 결과물
            print(str(data).replace('\'', '\"').replace('nan', '\"\"'))  # dic > str 형변환 & ' > ", nan > "" replace
            # http://json.parser.online.fr/ 에서 결과값 확인
            print('products.json 생성 완료!!')

        return 1
    except:
        return 0

def excel_2_json_hospitals(excel_path='hospitals_info.xlsx',json_path="./hospitals.json"):
    df = pd.read_excel(excel_path, index_col=None, engine='openpyxl')
    json_dic = {}
    df["병원 주소(위도)"] = df["병원 주소(위도)"].apply(str)
    df["병원 주소(경도)"] = df["병원 주소(경도)"].apply(str)
    df["위경도"] = list(df["병원 주소(위도)"] + "," + df["병원 주소(경도)"])
    print(df["위경도"])
    for i in range(len(df)):
        tmp_dic = {}
        tmp_dic["병원명"] = df["병원명"][i]
        tmp_dic["병원 주소"] = df["병원 주소"][i]
        tmp_dic["병원 홈페이지"] = df["병원 홈페이지"][i]
        tmp_dic["병원 전화번호"] = df["병원 전화번호"][i]
        json_dic[df["위경도"][i]] = tmp_dic
    with open(json_path, 'w', encoding='utf-8') as file:
        json.dump(json_dic, file)
    with open(json_path, 'r', encoding='utf-8') as file:
        json_data = json.load(file)
        print(json_data)
        print(json_data.keys())

    return "완료"

# 메인함수
if __name__ == '__main__':
    excel_2_json_hospitals()
    # # 제품 엑셀 데이터, json 파일로 생성하기
    # excel_2_json_products('추천_제품_리스트.xlsx')

    # with open('products.json', 'r') as f:
    #     json_data = json.load(f)
    # print('products json 파일>>>',json.dumps(json_data))
    # print('json 파일 요소 접근>>>',json_data['scalp_type']['양호']['브랜드평판지수'][0]['product_name'])

    # ------------- [ 네이버 지도에서 " 지역 + (모발, 두피, 탈모) + 병원" 검색 후 정보 추출 ] -----------------------
    # #search_result_2_json('모발') # 두피 검색 > 56 경기도 동안구 두피병원 진행. 404에러, 재시도중..
    #
    # # 엑셀 파일에 네이버 지도에서 검색 후 수집한 json데이터 저장
    # wb = Workbook()
    # ws = wb.active
    # ws.title = "병원정보"  # 엑셀 시트명 변경
    # ws.append(["검색어", "병원명", "병원 주소","병원 홈페이지","병원 전화번호","병원 주소(위도)","병원 주소(경도)"])
    #
    # for i, item in enumerate(os.listdir('C:\\Users\\hk_edu\\PycharmProjects\\datascraping\\data')):
    #     file_path = "data\\" + item
    #     print(file_path)
    #
    #     with open(file_path, 'r') as file:
    #         data = json.load(file)
    #         if data == '검색결과 없음':
    #             print("검색결과 없음!!")
    #             ws.append([item.replace('.json', '').lstrip('0123456789'), # 검색어
    #                        '결과 없음',                                             # 병원명
    #                        '결과 없음',                                             # 병원 주소
    #                        '결과 없음',                                             # 병원 홈페이지
    #                        '결과 없음',                                             # 병원 전화번호
    #                        '결과 없음',                                             # 병원 주소(위도)
    #                        '결과 없음'])                                            # 병원 주소(경도)
    #         else:
    #             json_data=json.loads(data)
    #
    #             # 위경도 값 계산
    #
    #             ws.append([item.replace('.json', '').lstrip('0123456789'),  # 검색어
    #                        json_data['site']['list'][0]['name'],            # 병원명
    #                        json_data['site']['list'][0]['address'],         # 병원 주소
    #                        json_data['site']['list'][0]['homePage'],        # 병원 홈페이지
    #                        json_data['site']['list'][0]['tel'],             # 병원 전화번호
    #                        get_lat_and_log(json_data['site']['list'][0]['address'])[0],     # 병원 주소(위도)
    #                        get_lat_and_log(json_data['site']['list'][0]['address'])[1]])    # 병원 주소(경도)
    #
    # wb.save("hospitals_info.xlsx")



#     ----- [ 대다모 개별 병원 정보 추출 후 엑셀에 저장] -------------------------------------------------------------
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "병원정보"  # 엑셀 시트명 변경
#     ws.append(["대다모 개별 병원 상세 링크", "병원명", "병원 주소", "병원 홈페이지", "병원 전화번호", "병원 주소(위도)", "병원 주소(경도)"])
#
#     for page in range(1,11):
#         res = [] # 결과값 저장
#         url = "https://daedamo.com/new/bbs/board.php?bo_table=hlist?&page=" + str(page)
#         html = requests.get(url).text
#         bs = BeautifulSoup(html, "html5lib")
#         for i_link in bs.select(".wr_id_first > a"):
#             #print(i_link['href'])
#             print(i_link['href'])
#             hos_info=dedamo_info(i_link['href']) # 대다모 개별 병원 정보 추출
#             res.extend( hos_info ) # 대다모 개별 병원 링크, 병원명 ~ 병원 전화번호 입력
#             #print('위치:',hos_info[2])
#             res.append( get_lat_and_log( hos_info[2] )[0]) # 위도 입력
#             res.append( get_lat_and_log( hos_info[2] )[1])  # 경도 입력
#
#
#             ws.append(res)
#
#             res=[]
#
#     wb.save("hospitals_info_dedamo.xlsx")
#     # https://daedamo.com/new/bbs/board.php?bo_table=hlist&wr_id=618&page=5
#     # dedamo_info('https://daedamo.com/new/bbs/board.php?bo_table=hlist&wr_id=618&page=5')
#     # dedamo_info('https://daedamo.com/new/bbs/board.php?bo_table=hlist&wr_id=616&page=5')




# 테스트
'''
    # url = "https://m.map.naver.com/search2/search.naver?query=" + keywords[6]
    # print(url)
    # html = requests.get(url).text
    # # print('>>',html)
    # print('html에서 json str 추출-------------------------------------')
    # html = html.split('var searchResult = ')[1]
    # html = html.split('var searchService =')[0]
    # html = html.replace(';', '').strip()
    # #html = '\'' + html + '\''
    # print(html)

    #print('m2-------------------------------------')
    # dict_object = json.loads(html)
    # print(dict_object)
    # print(type(dict_object))
    # print('len>>>>>',len(dict_object['site']['list']))
    # print('m3-------------------------------------')
    # bs = BeautifulSoup(html, "html5lib")
    # print(bs)
    # print(bs.select("._1Az1K"))
    # tags = bs.select(".search_list _items")
    # tag = tags[0]
    # print(tag.text)
    # while (1):
    #     pass
'''

# 네이버 지도 정보 크롤링 # https://velog.io/@toto9602/Selenium-%ED%99%9C%EC%9A%A9%ED%95%98%EC%97%AC-%EB%84%A4%EC%9D%B4%EB%B2%84-%EC%A7%80%EB%8F%84-%ED%81%AC%EB%A1%A4%EB%A7%81
'''
# ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ
# url = "https://map.naver.com/v5/search/탈모병원"# + itm
# html = requests.get(url).text
# print(html.encoding)
# print(html)
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC  # selenium에서 사용할 모듈 import
#
# import time
# import requests
# from bs4 import BeautifulSoup
# import re
# import csv
#
# driver = webdriver.Chrome("./103/chromedriver")  # selenium 사용에 필요한 chromedriver.exe 파일 경로 지정
#
# driver.get("https://map.naver.com/v5/")  # 네이버 신 지도
# try:
#     element = WebDriverWait(driver, 10).until(
#         EC.presence_of_element_located((By.CLASS_NAME, "input_search"))
#     )  # 입력창이 뜰 때까지 대기
# finally:
#     pass
# search_box = driver.find_element(By.CLASS_NAME("input_search")).send_keys("서울 칵테일바")  # 검색창에 "서울 칵테일바" 입력
# search_box.send_keys(Keys.ENTER)
# time.sleep(7)  # 화면 표시 기다리기
# frame = driver.find_element(By.CSS_SELECTOR("iframe#searchIframe"))
#
# driver.switch_to.frame(frame)
#
# time.sleep(3)
# # 여기까지 iframe 전환
#
# scroll_div = driver.find_element(By.XPATH("/html/body/div[3]/div/div[2]/div[1]"))
# # 검색 결과로 나타나는 scroll-bar 포함한 div 잡고
# driver.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
# time.sleep(2)
# driver.execute_script("arguments[0].scrollBy(0,2000);", scroll_div)
# time.sleep(2)
# driver.execute_script("arguments[0].scrollBy(0,2000);", scroll_div)
# time.sleep(2)
# driver.execute_script("arguments[0].scrollBy(0,2000);", scroll_div)
# time.sleep(2)
# driver.execute_script("arguments[0].scrollBy(0,2000);", scroll_div)
# time.sleep(2)
# # 여기까지 scroll
# # 맨 아래까지 내려서 해당 페이지의 내용이 다 표시되게 함
#
# # csv 파일 생성
# file = open('stores.csv', mode='w', newline='')
# writer = csv.writer(file)
# writer.writerow(["place", "rate", "address", "info", "image"])
# final_result = []
# time.sleep(1)
# # # 반복 시작
#
# i = 2
# while i <= 5:  # 몇 페이지까지 크롤링할 것인지 지정
#     stores_box = driver.find_element_by_xpath("/html/body/div[3]/div/div[2]/div[1]/ul")
#     stores = driver.find_elements_by_css_selector("li._3t81n._1l5Ut")
#     # 해당 페이지에서 표시된 모든 가게 정보
#
#     for store in stores:  # 한 페이지 내에서의 반복문. 순차적으로 가게 정보에 접근
#         name = store.find_element_by_css_selector("span._3Yilt").text  # 가게 이름
#         try:
#             rating = re.search('/span>(\d).', store.find_element_by_css_selector("span._3Yzhl._1ahw0").get_attribute(
#                 'innerHTML')).groups()[0]
#         except:
#             rating = ''
#         time.sleep(3)
#         # 평점 숫자 부분만 rating에 담음. 평점이 없는 경우가 있어 예외 처리
#         try:
#             img_src = re.search('url[(]"([\S]+)"',
#                                 store.find_element_by_css_selector("div.cb7hz.undefined").get_attribute(
#                                     'style')).groups()[0]
#         except:
#             img_src = ''
#         # 역시 대표 이미지가 없는 경우가 있어 예외 처리
#         click_name = store.find_element_by_css_selector("span._3Yilt")
#         click_name.click()
#         # 가게 주소, 홈페이지 링크를 확인하려면 가게 이름을 클릭해 세부 정보를 띄워야 함.
#
#         driver.switch_to.default_content()
#         time.sleep(7)
#         ##오래 헤맸던 부분!! switch_to.default_content()로 전환해야 frame_in iframe을 제대로 잡을 수 있다.
#
#         frame_in = driver.find_element_by_xpath(
#             '/html/body/app/layout/div[3]/div[2]/shrinkable-layout/div/app-base/search-layout/div[2]/entry-layout/entry-place-bridge/div/nm-external-frame-bridge/nm-iframe/iframe')
#
#         driver.switch_to.frame(frame_in)
#         # 가게 이름을 클릭하면 나오는 세부 정보 iframe으로 이동
#         time.sleep(3)
#         try:
#             address = re.search('서울\s(\w+)\s', driver.find_element_by_css_selector("span._2yqUQ").text).groups()[0]
#         except:
#             address = ''
#         # 주소 정보 확인
#         try:
#             link_url = driver.find_element_by_css_selector("a._1RUzg").text
#         except:
#             link_url = ''
#         # 홈페이지 url 확인
#         store_info = {
#             'placetitle': name,
#             'rate': rating,
#             'address': address,
#             'info': link_url,
#             'image': img_src
#         }
#         # 크롤링한 정보들을 store_info에 담고
#         print(name, rating, address, img_src, link_url)
#         print("*" * 50)
#         final_result.append(store_info)
#         # 출력해서 확인 후 final_result에 저장
#
#         driver.switch_to.default_content()
#         driver.switch_to.frame(frame)
#         time.sleep(8)
#         # 한 페이지 크롤링 끝
#
#     # '2'페이지로 이동하는 버튼 클릭 후 i 1증가
#     next_button = driver.find_element_by_link_text(str(i))
#     next_button.click()
#     i = i + 1
#     time.sleep(8)
#
# # while문이 종료되면 크롤링 종료
#
# for result in final_result:  # 크롤링한 가게 정보에 순차적으로 접근 & csv 파일 작성
#     row = []
#     row.append(result['placetitle'])
#     row.append(result['rate'])
#     row.append(result['address'])
#     row.append(result['info'])
#     row.append(result['image'])
#     writer.writerow(row)
#
# print(final_result)
# 최종 결과 확인
'''